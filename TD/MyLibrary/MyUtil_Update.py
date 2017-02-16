# uncompyle6 version 2.9.9
# Python bytecode 2.7 (62211)
# Decompiled from: Python 2.7.12 (v2.7.12:d33e0cf91556, Jun 27 2016, 15:24:40) [MSC v.1500 64 bit (AMD64)]
# Embedded file name: D:\robot\TD\MyLibrary\MyUtil_Update.py
# Compiled at: 2017-01-05 17:46:56
# -*- coding: utf-8 -*- 
import os
import xlrd
import xlwt
import openpyxl
import paramiko
import types
import datetime
import math
import string
import time
import sys
import zipfile
from pyPdf import PdfFileReader
from scp import SCPClient

def findUploadErrorMessageFile(path, upload_type):
    current_time = time.time()
    temp_dirs = [ rt for rt, dirs, files in os.walk(path) if abs(os.path.getmtime(rt) - current_time) < 60 and os.path.abspath(rt) != path ]
    target_file = ''
    if len(temp_dirs) > 0:
        for temp_path in temp_dirs:
            for temp_files in [ files for rt, dirs, files in os.walk(temp_path) ]:
                t = [ temp for temp in temp_files if temp.count('.txt') > 0 and (temp.count('Warnings of ' + upload_type.replace(' ', '')) > 0 or temp.count('Error of ' + upload_type.replace(' ', '')) > 0) ]
                if len(t) == 1:
                    target_file = temp_path + '\\' + t[0]
                elif len(t) > 1:
                    zipFile = [ tmp for tmp in temp_files if tmp.count('.zip') > 0 ]
                    zipName = zipFile[0].strip('\n').replace('.zip', '')
                    msgFile = [ tmp for tmp in temp_files if tmp.count(zipName) > 0 and tmp.count('.txt') > 0 ]
                    target_file = temp_path + '\\' + msgFile[0].strip('\n')

    else:
        raise Exception, 'get Error File Fail!'
    if len(target_file) > 0:
        return target_file
    print 'No Error/Warning File!'
    return target_file


def findUploadErrorDataFile(path, upload_type):
    current_time = time.time()
    temp_dirs = [ rt for rt, dirs, files in os.walk(path) if abs(os.path.getmtime(rt) - current_time) < 60 and os.path.abspath(rt) != path ]
    target_file = ''
    if len(temp_dirs) > 0:
        for temp_path in temp_dirs:
            for temp_files in [ files for rt, dirs, files in os.walk(temp_path) ]:
                t = [ temp for temp in temp_files if temp.count('Error data') > 0 ]
                if len(t) > 0:
                    target_file = temp_path + '\\' + t[0]

    else:
        raise Exception, 'get Error Data File Fail!'
    if len(target_file) > 0:
        return target_file
    print 'No Error Data File!'
    return target_file


def findUploadSourceFile(path, upload_type, sourceUploadFile):
    current_time = time.time()
    temp_dirs = [ rt for rt, dirs, files in os.walk(path) if abs(os.path.getmtime(rt) - current_time) < 60 and os.path.abspath(rt) != path ]
    target_file = ''
    if len(temp_dirs) > 0:
        for temp_path in temp_dirs:
            for temp_files in [ files for rt, dirs, files in os.walk(temp_path) ]:
                t = [ temp for temp in temp_files if temp.count(sourceUploadFile) > 0 ]
                if len(t) > 0:
                    target_file = temp_path + '\\' + t[0]

    else:
        raise Exception, 'get Source Upload File Fail!'
    if len(target_file) > 0:
        return target_file
    print 'No Source Upload File!'
    return target_file


def findLastFile(prod_path, prod_time):
    print prod_path
    print os.walk(prod_path)
    files = map(lambda t: prod_path + '\\' + t, [ pfiles for rt, dirs, pfiles in os.walk(prod_path) ][0])
    prod_files = [ f for f in files if os.path.getmtime(f) - prod_time >= 0 ]
    print prod_files
    if len(prod_files) == 1:
        return prod_files[0]
    raise Exception, 'Do not find!'


def getLinuxUploadErrorMessageFile(linux_ip, linux_user, linux_passwd, local_dir, test_user, upload_type):
    Host = linux_ip
    user = linux_user
    passwd = linux_passwd
    port = 22
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(Host, port, user, passwd)
    stdin, stdout, stderr = ssh.exec_command('find /usr/local/7thonline/7thOnline/data/temp -name *' + test_user + '* -mmin -1 -type d')
    dir = stdout.readlines()
    remotepath = ''
    target_file = ''
    for dir_tmp in dir:
        stdin, stdout, stderr = ssh.exec_command('ls ' + dir_tmp.strip('\n'))
        file_tmp = stdout.readlines()
        file = [ tmp for tmp in file_tmp if tmp.count('.txt') > 0 and (tmp.count('Warnings of ' + upload_type.replace(' ', '')) > 0 or tmp.count('Error of ' + upload_type.replace(' ', '')) > 0) ]
        if len(file) == 1:
            remotepath = dir_tmp.strip('\n') + '/' + file[0].strip('\n')
            target_file = file[0].strip('\n')
        elif len(file) > 1:
            zipFile = [ temp for temp in file_tmp if temp.count('.zip') > 0 ]
            zipName = zipFile[0].strip('\n').replace('.zip', '')
            msgFile = [ temp for temp in file_tmp if temp.count(zipName) > 0 and temp.count('.txt') > 0 ]
            remotepath = dir_tmp.strip('\n') + '/' + msgFile[0].strip('\n')
            target_file = msgFile[0].strip('\n')

    if len(remotepath) > 0:
        scpclient = SCPClient(ssh.get_transport(), socket_timeout=15.0)
        localpath = local_dir + target_file
        scpclient.get(remotepath, localpath)
        ssh.close()
        return localpath
    if len(dir) > 0:
        ssh.close()
        print 'No Error/Warning File!'
    else:
        ssh.close()
        raise Exception, 'get Error File Fail!'


def getLinuxUploadErrorDataFile(linux_ip, linux_user, linux_passwd, local_dir, test_user, upload_type):
    Host = linux_ip
    user = linux_user
    passwd = linux_passwd
    port = 22
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(Host, port, user, passwd)
    stdin, stdout, stderr = ssh.exec_command('find /usr/local/7thonline/7thOnline/data/temp -name *' + test_user + '* -mmin -1 -type d')
    dir = stdout.readlines()
    remotepath = ''
    target_file = ''
    for dir_tmp in dir:
        stdin, stdout, stderr = ssh.exec_command('ls ' + dir_tmp.strip('\n'))
        file_tmp = stdout.readlines()
        file = [ tmp for tmp in file_tmp if tmp.count('Error data') > 0 ]
        if len(file) > 0:
            remotepath = dir_tmp.strip('\n') + '/' + file[0].strip('\n')
            target_file = file[0].strip('\n')

    if len(remotepath) > 0:
        scpclient = SCPClient(ssh.get_transport(), socket_timeout=15.0)
        localpath = local_dir + target_file
        scpclient.get(remotepath, localpath)
        ssh.close()
        return localpath
    if len(dir) > 0:
        ssh.close()
        print 'No Error Data File!'
    else:
        ssh.close()
        raise Exception, 'get Error Data File Fail!'


def getLinuxUploadSourceFile(linux_ip, linux_user, linux_passwd, local_dir, test_user, upload_type, sourceUploadFile):
    Host = linux_ip
    user = linux_user
    passwd = linux_passwd
    port = 22
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(Host, port, user, passwd)
    stdin, stdout, stderr = ssh.exec_command('find /usr/local/7thonline/7thOnline/data/temp -name *' + test_user + '* -mmin -1 -type d')
    dir = stdout.readlines()
    remotepath = ''
    target_file = ''
    for dir_tmp in dir:
        stdin, stdout, stderr = ssh.exec_command('ls ' + dir_tmp.strip('\n'))
        file_tmp = stdout.readlines()
        file = [ tmp for tmp in file_tmp if tmp.count(sourceUploadFile) > 0 ]
        if len(file) > 0:
            remotepath = dir_tmp.strip('\n') + '/' + file[0].strip('\n')
            target_file = file[0].strip('\n')

    if len(remotepath) > 0:
        scpclient = SCPClient(ssh.get_transport(), socket_timeout=15.0)
        localpath = local_dir + target_file
        scpclient.get(remotepath, localpath)
        ssh.close()
        return localpath
    if len(dir) > 0:
        ssh.close()
        print 'No Source Upload File!'
    else:
        ssh.close()
        raise Exception, 'get Source Upload File Fail!'


def compareXLSX(prodf, stagef, diff_dir):  
    reload(sys)
    sys.setdefaultencoding('utf-8')
    error_flag = False
    dir = diff_dir.replace(".xlsx",".txt")
    error_info=""
    if(prodf == None or stagef == None ):
        error_flag = True
        error_info="Report isn't generated!"
    elif (prodf.find("xlsx") == -1 or stagef.find("xlsx") == -1 ) :
        error_flag = True
        error_info="Report isn't XLSX!"
    else:		
        pbook= openpyxl.load_workbook(prodf)
        sbook= openpyxl.load_workbook(stagef)	
        if (len(pbook.get_sheet_names()) != len(sbook.get_sheet_names()) ):
            error_flag = True
            error_info = error_info+ "tabs count diff \n" 	
        else:	
            for i  in range(0,len(sbook.get_sheet_names())):
                psh = pbook.worksheets[i]
                ssh = sbook.worksheets[i] 		
                maxrows = max(psh.max_row, ssh.max_row)			
                maxcols = max(psh.max_column, ssh.max_column)		
                if (psh.title == ssh.title):
                    for j in range(1, maxrows+1):	
                        row_diffCount = 0
                        for k in range(1, maxcols+1):	
                            if (psh.cell(row=j, column=k).value != ssh.cell(row=j, column=k).value ):
                                error_flag = True
                                row_diffCount +=1
                                if(row_diffCount == 1):
                                    error_info = error_info+("\n diff---%s---row%d" %(psh.title,j))
                                error_info = error_info+("---col%d[%s,%s]" %(k,ssh.cell(row=j, column=k).value,psh.cell(row=j, column=k).value ))
                else:
                    error_flag = True
                    error_info = error_info+("diff ---  %s\ %s; ---- rows:\  %s\%s;---- cols:\  %s\%s.\n" %(ssh.title, psh.title ,ssh.max_row,psh.max_row,ssh.max_column,psh.max_column))			 
                    for m in range(1, maxrows+1):		
                        row_diffCount = 0
                        for n in range(1, maxcols+1):	
                            if (psh.cell(row=m, column=n).value != ssh.cell(row=m, column=n).value ):
                                row_diffCount +=1;
                                if(row_diffCount == 1):
                                    error_info = error_info+("\n diff---row%d" %(m))
                                error_info = error_info+("---col%d[%s,%s]" %(n,ssh.cell(row=m, column=n).value ,psh.cell(row=m, column=n).value))		
    if(error_flag):
        diff_file = open(dir,"w")
        diff_file.write(error_info)	
        diff_file.close()
        raise Exception, "Excels are diff or not generated!"
    else:	
        print  "Same excels!"
    return


def copyZKVisibleXLSX(source_excel):
    new_excel = source_excel.replace('.xlsx', '_visible.xlsx', 1)
    source = openpyxl.load_workbook(source_excel)
    new = openpyxl.Workbook(encoding='utf-8')
    newTabsCount = 0
    for i in range(0, len(source.get_sheet_names())):
        if source.worksheets[i].title.find('_template') > 0:
            continue
        else:
            s_sheet = source.worksheets[i]
            n_sheet = new.create_sheet(index=newTabsCount, title=source.worksheets[i].title)
            colsCount = 1
            for j in range(1, s_sheet.max_column + 1):
                col_title = openpyxl.cell.get_column_letter(j)
                if s_sheet.column_dimensions[col_title].hidden:
                    continue
                for k in range(1, s_sheet.max_row + 1):
                    n_cell = '%s%d' % (openpyxl.cell.get_column_letter(colsCount), k)
                    s_cell = '%s%d' % (col_title, k)
                    n_sheet.cell(n_cell).value = s_sheet.cell(s_cell).value

                colsCount += 1

            newTabsCount += 1

    new.save(new_excel)
    return new_excel


def copyVisibleXLSX(source_excel, new_excel, tabs):
    tabNames = tabs.split(',')
    source = openpyxl.load_workbook(source_excel)
    new = openpyxl.Workbook(encoding='utf-8')
    for i in range(0, len(tabNames)):
        s_sheet = source.get_sheet_by_name(tabNames[i])
        if s_sheet:
            n_sheet = new.create_sheet(index=i, title=tabNames[i])
            colsCount = 1
            for j in range(1, s_sheet.max_column + 1):
                col_title = openpyxl.cell.get_column_letter(j)
                if s_sheet.column_dimensions[col_title].hidden:
                    continue
                for k in range(1, s_sheet.max_row + 1):
                    n_cell = '%s%d' % (openpyxl.cell.get_column_letter(colsCount), k)
                    s_cell = '%s%d' % (col_title, k)
                    n_sheet.cell(n_cell).value = s_sheet.cell(s_cell).value

                colsCount += 1

    new.save(new_excel)


def copyVisibleXLS(source_excel, new_excel, tabs):
    tabNames = tabs.split(',')
    source = xlrd.open_workbook(source_excel)
    new = xlwt.Workbook(encoding='utf-8')
    for i in range(0, len(tabNames)):
        s_sheet = source.sheet_by_name(tabNames[i])
        if s_sheet:
            n_sheet = new.add_sheet(tabNames[i])
            cinfo_map = s_sheet.colinfo_map
            colsCount = 0
            for j in range(0, s_sheet.ncols):
                if cinfo_map.get(j, 0) and cinfo_map[j].hidden == 1:
                    continue
                for k in range(0, len(s_sheet.col_values(j))):
                    n_sheet.write(k, colsCount, s_sheet.cell(k, j).value)
                    print '%s[row%d,col%d]' % (tabNames[i], k, colsCount)

                colsCount += 1

    new.save(new_excel)


def compareExcel(pf, sf, diff_dir):
    reload(sys)
    sys.setdefaultencoding('utf-8')
    error_flag = False
    dir = diff_dir.replace('.xlsx', '_diff.txt').replace('.xls', '_diff.txt')
    error_info = ''
    if pf == None or sf == None:
        error_flag = True
        error_info = "Report isn't generated!"
    else:
        pbook = xlrd.open_workbook(pf)
        sbook = xlrd.open_workbook(sf)
        if len(pbook.sheet_names()) != len(sbook.sheet_names()):
            error_flag = True
            error_info = error_info + 'tabs count or name diff \n'
        else:
            for i in range(0, len(pbook.sheet_names())):
                psh = pbook.sheet_by_index(i)
                ssh = sbook.sheet_by_index(i)
                minrows = min(psh.nrows, ssh.nrows)
                maxrows = max(psh.nrows, ssh.nrows)
                if psh.name == ssh.name and psh.nrows == ssh.nrows and psh.ncols == ssh.ncols:
                    for t in range(0, psh.nrows):
                        if psh.row_values(t) != ssh.row_values(t):
                            error_flag = True
                            error_info = error_info + '\n diff---%s---row%d' % (psh.name, t + 1)
                            if len(psh.row_values(t)) == len(ssh.row_values(t)):
                                for p in range(0, len(psh.row_values(t))):
                                    if psh.cell(t, p).value != ssh.cell(t, p).value:
                                        error_info = error_info + '---col%d[%s,%s]' % (p + 1, ssh.cell(t, p).value, psh.cell(t, p).value)

                            else:
                                for p in range(0, max(len(psh.row_values(t)), len(ssh.row_values(t)))):
                                    if p < min(len(psh.row_values(t)), len(ssh.row_values(t))) and psh.cell(t, p).value != ssh.cell(t, p).value:
                                        error_info = error_info + '---col%d[%s,%s]' % (p + 1, ssh.cell(t, p).value, psh.cell(t, p).value)
                                    if p >= min(len(psh.row_values(t)), len(ssh.row_values(t))) and len(psh.row_values(t)) > len(ssh.row_values(t)):
                                        error_info = error_info + '---col%d[ ,%s]' % (p + 1, psh.cell(t, p).value)
                                    if p >= min(len(psh.row_values(t)), len(ssh.row_values(t))) and len(psh.row_values(t)) < len(ssh.row_values(t)):
                                        error_info = error_info + '---col%d[%s, ]' % (p + 1, ssh.cell(t, p).value)

                else:
                    error_flag = True
                    error_info = error_info + 'diff ---  %s\\ %s; ---- rows:\\  %s\\%s;---- cols:\\  %s\\%s.\n' % (ssh.name, psh.name, ssh.nrows, psh.nrows, ssh.ncols, psh.ncols)
                    for j in range(0, maxrows):
                        if j < minrows and psh.row_values(j) != ssh.row_values(j):
                            error_info = error_info + '\n diff---%s\\ %s---row%d' % (ssh.name, psh.name, j + 1)
                            for k in range(0, max(len(psh.row_values(j)), len(ssh.row_values(j)))):
                                if k < min(len(psh.row_values(j)), len(ssh.row_values(j))) and psh.cell(j, k).value != ssh.cell(j, k).value:
                                    error_info = error_info + '---col%d[%s,%s]' % (k + 1, ssh.cell(j, k).value, psh.cell(j, k).value)
                                if k >= min(len(psh.row_values(j)), len(ssh.row_values(j))) and len(psh.row_values(j)) > len(ssh.row_values(j)):
                                    error_info = error_info + '---col%d[ ,%s]' % (k + 1, psh.cell(j, k).value)
                                if k >= min(len(psh.row_values(j)), len(ssh.row_values(j))) and len(psh.row_values(j)) < len(ssh.row_values(j)):
                                    error_info = error_info + '---col%d[%s, ]' % (k + 1, ssh.cell(j, k).value)

                        if j >= minrows and psh.nrows == maxrows:
                            error_info = error_info + '\n diff---%s---row%d is not in stage.' % (psh.name, j + 1)
                        if j >= minrows and ssh.nrows == maxrows:
                            error_info = error_info + '\n diff---%s---row%d is not in prod.' % (ssh.name, j + 1)

    if error_flag:
        diff_file = open(dir, 'w')
        diff_file.write(error_info)
        diff_file.close()
        raise Exception, 'Excels are diff or not generated!'
    else:
        return 'Same excels!'
    return


def getEnvironmentSetting(fpath, tabName, envRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    env = [sheet.cell(0, 1).value]
    for p in range(1, envRows):
        env.append(sheet.cell(p, 1).value)

    return env


def getUsersCount(fpath, tabName, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    count = sheet.nrows - int(tilteRows)
    return count


def getUserCriteria(fpath, tabName, user_number, startCol, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if user_number >= sheet.nrows - int(tilteRows) or user_number < 0:
        raise Exception, error_message + 'user does not existing!'
    else:
        criteria = sheet.row_values(user_number + int(tilteRows))
        del criteria[0:int(startCol)]
        return [ criteria for criteria in criteria if criteria.strip() != '' ]


def getUser(fpath, tabName, user_number, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if user_number >= sheet.nrows - int(tilteRows) or user_number < 0:
        raise Exception, error_message + 'user does not existing!'
    else:
        return sheet.cell(user_number + int(tilteRows), 0).value


def getCompanyInfo(fpath, tabName, user_number, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if user_number >= sheet.nrows - int(tilteRows) or user_number < 0:
        raise Exception, error_message + 'user does not existing!'
    else:
        info = [
         str(sheet.cell(user_number + int(tilteRows), 1).value)]
        info.append(str(sheet.cell(user_number + int(tilteRows), 2).value))
        return info


def getYears(fpath, tabName, user_number, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if user_number >= sheet.nrows - int(tilteRows) or user_number < 0:
        raise Exception, error_message + 'user does not existing!'
    else:
        year = str(sheet.cell(user_number + int(tilteRows), 6).value).replace('.0', '').split(',')
        return year


def getMonths(fpath, tabName, user_number, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if user_number >= sheet.nrows - int(tilteRows) or user_number < 0:
        raise Exception, error_message + 'user does not existing!'
    else:
        month = str(sheet.cell(user_number + int(tilteRows), 7).value).replace('.0', '').split(',')
        return month


def getDownloadType(fpath, tabName, user_number, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if user_number >= sheet.nrows - int(tilteRows) or user_number < 0:
        raise Exception, error_message + 'user does not existing!'
    else:
        return sheet.cell(user_number + int(tilteRows), 3).value


def getDownloadOption(fpath, tabName, user_number, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if user_number >= sheet.nrows - int(tilteRows) or user_number < 0:
        raise Exception, error_message + 'user does not existing!'
    else:
        return sheet.cell(user_number + int(tilteRows), 4).value


def getCellValues(fpath, tabName, user_number, col_number, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if int(user_number) >= sheet.nrows - int(tilteRows) or int(user_number) < 0:
        raise Exception, error_message + 'user does not existing!'
    else:
        return str(sheet.cell(int(user_number) + int(tilteRows), int(col_number)).value).strip().split(',')


def getCellValuesListByRowCol(fpath, tabName, row_number, col_number):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if int(row_number) < 1 or int(col_number) < 1:
        raise Exception, error_message + 'cell does not existing in sheet!'
    else:
        return str(sheet.cell(int(row_number) - 1, int(col_number) - 1).value).strip().split(',')


def getCellValueByRowCol(fpath, tabName, row_number, col_number):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if int(row_number) < 1 or int(col_number) < 1:
        raise Exception, error_message + 'cell does not existing in sheet!'
    else:
        return str(sheet.cell(int(row_number) - 1, int(col_number) - 1).value).strip()


def getWSIDList(WSID_string, replace_str):
    WSid = WSID_string.replace(replace_str, '').split(',')
    return [ WSid for WSid in WSid if WSid.strip() != '' ]


def getListSize(options):
    return len(str(options).split(','))


def diffText(stage_text, prod_text):
    if stage_text == prod_text:
        return False
    else:
        return True


def comparePDF(pf, sf):
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    error_flag = False
    error_info = ''
    if pf == None and sf == None:
        print "Prod and Stage WS PDF DON'T EXIST!"
    elif pf != None and sf == None:
        error_flag = True
        error_info = error_info + "Stage WS PDF DOESN'T EXIST!"
    elif pf == None and sf != None:
        error_flag = True
        error_info = error_info + "Prod WS PDF DOESN'T EXIST!"
    else:
        pPdf = PdfFileReader(file(pf, 'rb'))
        sPdf = PdfFileReader(file(sf, 'rb'))
        if pPdf.getNumPages() != sPdf.getNumPages():
            error_flag = True
            error_info = error_info + 'Diff WS PDF Page Numbers! ---stage pages: %d---prod pages: %d \n' % (sPdf.getNumPages(), pPdf.getNumPages())
        else:
            for p in range(0, sPdf.getNumPages()):
                ptext = pPdf.getPage(p).extractText()
                stext = sPdf.getPage(p).extractText()
                ptag = ptext[ptext.find('[') + 1:ptext.find(']')]
                stag = stext[stext.find('[') + 1:stext.find(']')]
                ptag1 = ptext[ptext.find('Createdby') - 23:ptext.find('Createdby')]
                stag1 = stext[stext.find('Createdby') - 23:stext.find('Createdby')]
                ptxt = ptext.replace(ptag, '').replace(ptag1, '')
                stxt = stext.replace(stag, '').replace(stag1, '')
                if ptxt != stxt:
                    error_flag = True
                    error_info = error_info + 'Diff WS PDF Page %d! \n' % (p + 1)

    if error_flag:
        raise Exception, error_info
    else:
        print 'Same WS PDF!'
    return


def extractZIP(pf, path, prefix):
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    z = zipfile.ZipFile(pf, 'r')
    fileNameList = []
    for f in z.namelist():
        fileName = path + prefix + f
        output = open(fileName, 'wb')
        output.write(z.read(f))
        output.close()
        fileNameList.append(fileName)

    return fileNameList


def compareExcelWithXls(pf, sf):
    error_flag = False
    error_info = ''
    if pf == None or sf == None:
        error_flag = True
        error_info = "Report isn't generated!"
    else:
        pbook = xlrd.open_workbook(pf, formatting_info=True)
        sbook = xlrd.open_workbook(sf, formatting_info=True)
        if len(pbook.sheet_names()) != len(sbook.sheet_names()):
            error_flag = True
            error_info = error_info + 'tab name diff \n'
        else:
            for i in range(0, len(pbook.sheet_names())):
                psh = pbook.sheet_by_index(i)
                ssh = sbook.sheet_by_index(i)
                minrows = min(psh.nrows, ssh.nrows)
                minclos = min(psh.ncols, ssh.ncols)
                if psh.name == ssh.name and psh.nrows == ssh.nrows and psh.ncols == ssh.ncols:
                    for ct in range(0, psh.ncols):
                        if psh.colinfo_map[ct].hidden != ssh.colinfo_map[ct].hidden:
                            error_flag = True
                            error_info = error_info + '\n Hidden Diff---%s---col%d' % (psh.name, ct + 1)

                    for t in range(0, psh.nrows):
                        if psh.rowinfo_map[t].hidden != ssh.rowinfo_map[t].hidden:
                            error_flag = True
                            error_info = error_info + '\n Hidden Diff---%s---row%d' % (psh.name, t + 1)
                        if psh.row_values(t) != ssh.row_values(t):
                            error_flag = True
                            error_info = error_info + '\n diff---%s---row%d' % (psh.name, t + 1)
                            for p in range(0, psh.ncols):
                                if psh.cell(t, p).value != ssh.cell(t, p).value:
                                    error_info = error_info + '---col%d' % (p + 1)

                else:
                    error_flag = True
                    error_info = error_info + 'diff ---  %s\\ %s; ---- rows:\\  %s\\%s;---- cols:\\  %s\\%s.\n' % (psh.name, ssh.name, psh.nrows, ssh.nrows, psh.ncols, ssh.ncols)
                    for cj in range(0, minclos):
                        if psh.colinfo_map[cj].hidden != ssh.colinfo_map[cj].hidden:
                            error_info = error_info + '\n Hidden Diff---%s---col%d' % (psh.name, cj + 1)

                    for j in range(0, minrows):
                        if psh.rowinfo_map[j].hidden != ssh.rowinfo_map[j].hidden:
                            error_info = error_info + '\n Hidden Diff---%s---row%d' % (psh.name, j + 1)
                        if psh.row_values(j) != ssh.row_values(j):
                            error_info = error_info + '\n diff---%s---row%d' % (psh.name, j + 1)
                            for k in range(0, minclos):
                                if psh.cell(j, k).value != ssh.cell(j, k).value:
                                    error_info = error_info + '---col%d' % (k + 1)

    if error_flag:
        raise Exception, error_info
    else:
        print 'Same excels!'
    return


def getCellValue(fpath, tabName, user_number, col_number, tilteRows):
    book = xlrd.open_workbook(fpath)
    sheet = book.sheet_by_name(tabName)
    if int(user_number) >= sheet.nrows - int(tilteRows) or int(user_number) < 0:
        raise Exception, error_message + 'user does not existing!'
    else:
        return sheet.cell(int(user_number) + int(tilteRows), int(col_number)).value


def writeZKLoadTestingParamForDT(userName, passwd, receiver, editCell, file_dir):
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    dir = file_dir.replace('/', ' ')
    config_head = 'userName,passwd,reciver_0,reciver_0_Y,reciver_1,reciver_1_Y,reciver_2,reciver_2_Y,reciver_3,reciver_3_Y,reciver_4,reciver_4_Y,reciver_5,reciver_5_Y,reciver_6,reciver_6_Y,reciver_7,reciver_7_Y,reciver_8,reciver_8_Y,reciver_9,reciver_9_Y,selectedReceiver,openStockCol,openStockRow,openStockNum,openStockX,openStockY,prepackTotalCol,prepackTotalRow,prepackTotalNum,prepackTotalX,prepackTotalY\n'
    config_file = open(dir, 'a+')
    config_data = config_file.read()
    if len(receiver) > 0:
        userInfo = userName + ',' + passwd + ','
        receiverList = receiver.split(';')
        for i in range(0, 10):
            if i < len(receiverList):
                userInfo += receiverList[i] + ','
            else:
                userInfo += ' ,'

        userInfo += editCell + '\n'
    else:
        userInfo = userName + ',' + passwd + ',' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ',' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + ' ,' + editCell + '\n'
    if len(config_data) == 0:
        config_file.write(config_head)
    config_file.seek(0, 2)
    config_file.write(userInfo)
    config_file.close()


def writeZKLoadTestingParamForTransfer(userName, passwd, title, styleCount, editBtnRowInfo, modifyRow, addDoorRowInfo, fromDoorInfo, toDoorInfo, fromDoorSize, file_dir):
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    dir = file_dir.replace('/', ' ')
    config_head = title + '\n'
    config_file = open(dir, 'a+')
    config_data = config_file.read()
    if len(fromDoorInfo) > 0:
        userInfo = userName + ',' + passwd + ',' + styleCount + ',' + editBtnRowInfo + ',' + addDoorRowInfo + ',' + fromDoorInfo + ',' + toDoorInfo + ',' + modifyRow + ',' + fromDoorSize + '\n'
    if len(config_data) == 0:
        config_file.write(config_head)
    config_file.seek(0, 2)
    config_file.write(userInfo)
    config_file.close()


def saveDBQueryDataToFile(filePath, data, fileType='', delimete=''):
    os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    if delimete:
        config_file = open(filePath, 'a+')
        config_data = config_file.read()
        for row in data:
            config_file.seek(0, 2)
            for item in row:
                if item == None:
                    item = ''
                config_file.write(str(item) + delimete)

            config_file.write('\n')

        config_file.close()
    else:
        w = xlwt.Workbook()
        ws = w.add_sheet('Sheet1')
        for row_num in range(0, len(data)):
            row = list(data[row_num])
            for col_num in range(0, len(row)):
                if type(row[col_num]) is types.StringType:
                    row[col_num] = row[col_num].decode('utf-8')
                if row[col_num] != None:
                    ws.write(row_num, col_num, row[col_num])

        w.save(filePath)
    return


def comprareStringToSortByAscending(str1, str2):
    strList = [str1, str2]
    strList.sort()
    if str1 == strList[0]:
        flag = True
    else:
        flag = False
    return flag


def mergeZKColData(dataList, topHead, dataList_new, topHead_new):
    startCol = topHead_new.index(topHead[-1]) + 1
    if len(dataList) != len(dataList_new):
        raise Exception, 'Rows Count Diff !'
    else:
        for row_num in range(0, len(dataList)):
            for col_num in range(startCol, len(topHead_new)):
                topHead.append(topHead_new[col_num])
                dataList[row_num].append(dataList_new[row_num][col_num + 1])

        return (
         dataList, topHead)


def mergeZKRowData(list1=[], list2=[]):
    if list1[-1] in list2:
        startRow = list2.index(list1[-1]) + 1
    else:
        startRow = 0
    for row_num in range(startRow, len(list2)):
        list1.append(list2[row_num])

    return list1


def compareZKAppletWithExcel(ZKData, ZKTabName, downloadFile, diff_dir, titleRowsCount, expandDoorStartCol=100000, expandDoorFlag=False):
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    expandDoorFlag = bool(expandDoorFlag)
    error_flag = False
    dir = diff_dir
    error_info = ''
    if ZKData == None or downloadFile == None:
        error_flag = True
        error_info = "Report isn't generated!"
    else:
        pbook = xlrd.open_workbook(downloadFile)
        if len([ tmp for tmp in pbook.sheet_names() if tmp == ZKTabName ]) != 1:
            error_flag = True
            if len([ tmp for tmp in pbook.sheet_names() if tmp == ZKTabName ]) == 0:
                error_info = error_info + "Tab %s Isn't existing in download File!\n" % ZKTabName
            if len([ tmp for tmp in pbook.sheet_names() if tmp == ZKTabName ]) > 1:
                error_info = error_info + 'There are multiple %s tabs in download File!\n' % ZKTabName
        else:
            psh = pbook.sheet_by_name(ZKTabName)
            minrows = min(psh.nrows, len(ZKData))
            maxrows = max(psh.nrows, len(ZKData))
            if psh.nrows == len(ZKData):
                for t in range(0, psh.nrows):
                    excelRow = psh.row_values(t)
                    ZKRow = ZKData[t][1:]
                    for n in range(0, len(excelRow)):
                        if isinstance(excelRow[n], float):
                            if excelRow[n] == int(excelRow[n]):
                                excelRow[n] = str(int(excelRow[n]))
                            else:
                                excelRow[n] = str(excelRow[n])
                        if t < int(titleRowsCount) - 2 and n < int(expandDoorStartCol) - 1:
                            excelRow[n] = str(excelRow[n]).strip()
                        if (t == 1 or t == int(titleRowsCount) - 2) and n >= int(expandDoorStartCol) - 1 and int(titleRowsCount) != 2 and expandDoorFlag:
                            excelRow[n] = str(excelRow[n])[:4]
                        if t >= int(titleRowsCount) and n == 0 and str(excelRow[n].find('Total ')) != -1:
                            excelRow[n] = str(excelRow[n]).strip()
                        if t == int(titleRowsCount) - 1 and n >= int(expandDoorStartCol) - 1:
                            excelRow[n] = str(excelRow[n]).strip().replace('\n', '')

                    for m in range(0, len(ZKRow)):
                        if isinstance(ZKRow[m], float):
                            if ZKData[m] >= 0:
                                if ZKRow[m] == int(ZKRow[m]):
                                    ZKRow[m] = str(int(ZKRow[m]))
                                else:
                                    ZKRow[m] = str(ZKRow[m])
                        if ZKRow[m].find(u'\xa5') != -1 or ZKRow[m].find('$') != -1:
                            ZKRow[m] = str(int(float(ZKRow[m].replace(',','')[1:])))
                        if expandDoorFlag and t < int(titleRowsCount) - 2 and m < int(expandDoorStartCol) - 1 and m > 0:
                            ZKRow[m] = ''
                        if expandDoorFlag and (t == 1 or t == int(titleRowsCount) - 2) and m >= int(expandDoorStartCol) - 1 and int(titleRowsCount) != 2:
                            ZKRow[m] = str(ZKRow[m])[-4:]
                        if t >= int(titleRowsCount) and len([ rt for rt in ZKData[t] if rt.find('Total ') != -1 ]) >= 2 and ZKRow[m] == ZKRow[m - 1] and ZKRow[m].find('Total ') != -1:
                            for c in range(0, len([ rt for rt in ZKData[t] if rt.find('Total ') != -1 ]) - 1):
                                ZKRow[m + c] = ''

                    if excelRow != ZKRow and (t < int(titleRowsCount) - 5 or t > int(titleRowsCount) - 3):
                        error_flag = True
                        error_info = error_info + '\n diff---%s---row%d' % (psh.name, t + 1)
                        if len(psh.row_values(t)) == len(ZKData[t][1:]):
                            for p in range(0, len(psh.row_values(t))):
                                if excelRow[p] != ZKRow[p]:
                                    error_info = error_info + '---col%d[%s,%s]' % (p + 1, excelRow[p], ZKRow[p])

                        else:
                            for p in range(0, max(len(psh.row_values(t)), len(ZKData[t][1:]))):
                                if p < min(len(psh.row_values(t)), len(ZKData[t][1:])) and excelRow[p] != ZKRow[p]:
                                    error_info = error_info + '---col%d[%s,%s]' % (p + 1, excelRow[p], ZKRow[p])
                                if p >= min(len(psh.row_values(t)), len(ZKData[t][1:])) and len(psh.row_values(t)) > len(ZKData[t][1:]):
                                    error_info = error_info + '---col%d[%s, ]' % (p + 1, excelRow[p])
                                if p >= min(len(psh.row_values(t)), len(ZKData[t][1:])) and len(psh.row_values(t)) < len(ZKData[t][1:]):
                                    error_info = error_info + '---col%d[ ,%s]' % (p + 1, ZKRow[p])
            else:
                error_flag = True
                error_info = error_info + 'diff ---- rows:\\  %s\\%s .\n' % (psh.nrows, len(ZKData))
                for j in range(0, maxrows):
                    if (j < psh.nrows):
                        excelRow = psh.row_values(j)
                    else:
                        excelRow = []
                    if (j < len(ZKData)):
                        ZKRow = ZKData[j][1:]
                    else:
                        ZKRow = []
                    for a in range(0, len(excelRow)):
                        if isinstance(excelRow[a], float):
                            if excelRow[a] == int(excelRow[a]):
                                excelRow[a] = str(int(excelRow[a]))
                            else:
                                excelRow[a] = str(excelRow[a])
                        if j < int(titleRowsCount) - 2 and a < int(expandDoorStartCol) - 1:
                            excelRow[a] = str(excelRow[a]).strip()
                        if (j == 1 or j == int(titleRowsCount) - 2) and a >= int(expandDoorStartCol) - 1 and int(titleRowsCount) != 2 and expandDoorFlag:
                            excelRow[a] = str(excelRow[a])[:4]
                        if j >= int(titleRowsCount) and a == 0 and str(excelRow[a].find('Total ')) != -1:
                            excelRow[a] = str(excelRow[a]).strip()
                        if j == int(titleRowsCount) - 1 and a >= int(expandDoorStartCol) - 1:
                            excelRow[a] = str(excelRow[a]).strip().replace('\n', '')
                    for b in range(0, len(ZKRow)):
                        if isinstance(ZKRow[b], float):
                            if ZKRow[b] >= 0:
                                if ZKRow[b] == int(ZKRow[b]):
                                    ZKRow[b] = str(int(ZKRow[b]))
                                else:
                                    ZKRow[b] = str(ZKRow[b])
                        if ZKRow[b].find(u'\xa5') != -1 or ZKRow[b].find('$') != -1:
                            ZKRow[b] = str(int(float(ZKRow[b].replace(',','')[1:])))
                        if j < int(titleRowsCount) - 2 and b < int(expandDoorStartCol) - 1 and b > 0 and expandDoorFlag:
                            ZKRow[b] = ''
                        if (j == 1 or j == int(titleRowsCount) - 2) and b >= int(expandDoorStartCol) - 1 and int(titleRowsCount) != 2 and expandDoorFlag:
                            ZKRow[b] = str(ZKRow[b])[-4:]
                        if j >= int(titleRowsCount) and len([ rt for rt in ZKData[j] if rt.find('Total ') != -1 ]) >= 2 and ZKRow[b] == ZKRow[b - 1] and ZKRow[b].find('Total ') != -1:
                            for d in range(0, len([ rt for rt in ZKData[j] if rt.find('Total ') != -1 ]) - 1):
                                ZKRow[b + d] = ''

                    if j < minrows and excelRow != ZKRow and (j < int(titleRowsCount) - 5 or j > int(titleRowsCount) - 3):
                        error_info = error_info + '\n diff ---- row%d' % (j + 1)
                        for k in range(0, max(len(psh.row_values(j)), len(ZKData[j][1:]))):
                            if k < min(len(psh.row_values(j)), len(ZKData[j][1:])) and excelRow[k] != ZKRow[k]:
                                error_info = error_info + '---col%d[%s,%s]' % (k + 1, excelRow[k], ZKRow[k])
                            if k >= min(len(psh.row_values(j)), len(ZKData[j][1:])) and len(psh.row_values(j)) > len(ZKData[j][1:]):
                                error_info = error_info + '---col%d[%s, ]' % (k + 1, excelRow[k])
                            if k >= min(len(psh.row_values(j)), len(ZKData[j][1:])) and len(psh.row_values(j)) < len(ZKData[j][1:]):
                                error_info = error_info + '---col%d[ ,%s]' % (k + 1, ZKRow[k])

                    if j >= minrows and psh.nrows == maxrows:
                        error_info = error_info + '\n diff---row%d is not in ZK Applet.' % (j + 1)
                    if j >= minrows and len(ZKData) == maxrows:
                        error_info = error_info + '\n diff---row%d is not in Download File.' % (j + 1)

    if error_flag:
        diff_file = open(dir, 'w')
        diff_file.write(error_info)
        diff_file.close()
        raise Exception, 'Download excels are Difference from ZK Applet! %s' % expandDoorFlag
    else:
        return 'Download excels are same as ZK Applet !'
    return


def delZKBlankCols(ZKData, startCol):
    col = int(startCol) - 1
    colsCount = len(zip(*ZKData))
    if any(ZKData) and len(zip(*ZKData)) >= int(startCol):
        for rowNum in range(0, len(ZKData)):
            if col == int(startCol) - 1:
                for colNum in range(int(startCol) - 1, len(ZKData[rowNum])):
                    if not any([ t[colNum] for t in ZKData ]):
                        col = colNum
                        break

            if col != int(startCol) - 1:
                for num in range(col, colsCount):
                    del ZKData[rowNum][-1]

    elif not any(ZKData):
        raise Exception, 'ZK Data is Blank!'
    return ZKData


def copyZKXLSXWithoutExpandDoorInfo(source_excel, expandDoorStartCol, doorRowsCount):
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    new_excel = source_excel.replace('.xlsx', '_visible.xlsx', 1)
    source = openpyxl.load_workbook(source_excel)
    sbook = xlrd.open_workbook(source_excel)
    new = openpyxl.Workbook()
    newTabsCount = 0
    for i in range(0, len(source.get_sheet_names())):
        if source.worksheets[i].title.find('_template') > 0:
            continue
        else:
            s_sheet = source.worksheets[i]
            n_sheet = new.create_sheet(index=newTabsCount, title=source.worksheets[i].title)
            colsCount = 1
            for j in range(1, s_sheet.max_column + 1):
                col_title = openpyxl.utils.get_column_letter(j)
                if j > int(expandDoorStartCol) and (s_sheet.column_dimensions[col_title].hidden or sbook.sheet_by_index(i).col_values(j - 2, 0, int(doorRowsCount)) == sbook.sheet_by_index(i).col_values(j - 1, 0, int(doorRowsCount))):
                    continue
                for k in range(1, s_sheet.max_row + 1):
                    n_cell = '%s%d' % (openpyxl.utils.get_column_letter(colsCount), k)
                    s_cell = '%s%d' % (col_title, k)
                    if isinstance(s_sheet.cell(s_cell).value, str):
                        n_sheet.cell(n_cell).value = s_sheet.cell(s_cell).value.strip()
                    else:
                        n_sheet.cell(n_cell).value = s_sheet.cell(s_cell).value

                colsCount += 1

            newTabsCount += 1

    new.save(new_excel)
    return new_excel


def getColDataFromList(startRow, colNum, dataList):
    return [ t[int(colNum)] for t in dataList[int(startRow):] ]


def convertStrToDateInList(strList, dateFormat='%m/%d/%Y'):
    for a, b in enumerate(strList):
        strList[a] = datetime.datetime.strptime(b, dateFormat)

    return strList


def compareDateStrList(dateStrList, fromDateStr, toDateStr, titleEndRow=0, totalRowsList=[], dateFormat='%m/%d/%Y'):
    fromDate = datetime.datetime.strptime(fromDateStr, dateFormat)
    toDate = datetime.datetime.strptime(toDateStr, dateFormat)
    return [ (a + titleEndRow+1, b) for a, b in enumerate(dateStrList) if (a + titleEndRow) not in totalRowsList and (datetime.datetime.strptime(b, dateFormat) > toDate or datetime.datetime.strptime(b, dateFormat) < fromDate or not b)]


if __name__ == '__main__':
    findUploadErrorMessageFile()
    compareExcel()
    findLastFile()
    getEnvironmentsetting()
    getUsersCount()
    getUserCriteria()
    getUser()
    getCompanyInfo()
    getYears()
    getMonths()
    getDownloadType()
    getDownloadOption()
    getWSIDList()
    diffText()
    comparePDF()
    getCellValues()
    getListSize()
    extractZIP()
    compareExcelWithXls()
    getCellValue()
    copyVisibleXLS()
    copyVisibleXLSX()
    compareXLSX()
    copyZKVisibleXLSX()
    writeZKLoadTestingParamForDT()
    writeZKLoadTestingParamForTransfer()
    comprareStringToSortByAscending()
    getLinuxUploadErrorMessageFile()
    getCellValuesListByRowCol()
    getCellValueByRowCol()
    saveDBQueryDataToFile()
    getLinuxUploadErrorDataFile()
    findUploadErrorDataFile()
    getLinuxUploadSourceFile()
    findUploadSourceFile()
    mergeZKColData()
    mergeZKRowData()
    compareZKAppletWithExcel()
    delZKBlankCols()
    copyZKXLSXWithoutExpandDoorInfo()
    getColDataFromList()
    convertStrToDateInList()
    compareDateStrList()
# okay decompiling MyUtil_Update.pyc
